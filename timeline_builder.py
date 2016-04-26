#!/usr/bin/env/python
# 
#

from jinja2 import Environment
from tkinter.filedialog import askopenfilename
import openpyxl

htmlStart = """
<!DOCTYPE html>
<html>
<body>

<h2>
{{person}}
</h2> 
"""

htmlIndexStart = """
<!DOCTYPE html>
<html>
<body>

<h2>
{{familyName}} Family
</h2> 
"""

htmlIndex = """
<p>
<a href="{{personLink}}">
{{person}}
</a>
</p>
"""


htmlTimeline = """
<p style = "margin:0"	>
<a href="{{eventLink}}">
<img src="line.png" alt="event:" width="42" height="60" style="vertical-align:middle">
</a>
{{event}}
</p>
"""

htmlEnd = """
</body>
</html>
"""

htmlEvent = """
<h1>
{{event}}
</h1>
"""

htmlDate = """
<p>
When: {{month}}/{{day}}/{{year}}
</p>
"""

htmlWhere = """
<p>
Where: {{city}},{{state}}
</p>

"""

htmlWhereOther = """

<p>
Detailed Where: {{other}}
</p>

"""

htmlDescription = """

<p>
Description: {{shortDescription}}
</p>

"""

htmlPhoto = """

<img src={{photo}} alt="photo" style="width:304px;">

"""


def start_index_doc(FamilyName):
    #print(Environment().from_string(htmlStart).render(person=PersonName))
    return(Environment().from_string(htmlIndexStart).render(familyName=FamilyName))

def index_doc(Person):
    #print(Environment().from_string(htmlStart).render(person=PersonName))
    return(Environment().from_string(htmlIndex).render(person=Person, personLink=Person+'.html'))

def start_doc(PersonName):
    #print(Environment().from_string(htmlStart).render(person=PersonName))
    return(Environment().from_string(htmlStart).render(person=PersonName))

#------------------------------------------------------------------------------
    
def body_doc(Event, EventLink):
      #print(Environment().from_string(htmlTimeline).render(event=Event, eventLink=EventLink)) 
      return(Environment().from_string(htmlTimeline).render(event=Event, eventLink=EventLink))

#------------------------------------------------------------------------------
    
def end_doc():
    #print(Environment().from_string(htmlEnd).render())
    return(Environment().from_string(htmlEnd).render())

#------------------------------------------------------------------------------

def event_doc(Event):
    #print(Environment().from_string(htmlEnd).render())
    return(Environment().from_string(htmlEvent).render(event=Event))

#------------------------------------------------------------------------------

def date_doc(Day, Month, Year):
    #print(Environment().from_string(htmlEnd).render())
    return(Environment().from_string(htmlDate).render(day=Day, month=Month, year=Year))

#------------------------------------------------------------------------------

def where_doc(City, State):
    #print(Environment().from_string(htmlEnd).render())
    return(Environment().from_string(htmlWhere).render(city=City, state=State))

#------------------------------------------------------------------------------

def where_other_doc(Other):
    #print(Environment().from_string(htmlEnd).render())
    return(Environment().from_string(htmlWhereOther).render(other=Other))

#------------------------------------------------------------------------------

def description_doc(Decription):
    #print(Environment().from_string(htmlEnd).render())
    return(Environment().from_string(htmlDescription).render(shortDescription=Decription))

#------------------------------------------------------------------------------

def photo_doc(Photo):
    #print(Environment().from_string(htmlEnd).render())
    return(Environment().from_string(htmlPhoto).render(photo=Photo))

#------------------------------------------------------------------------------

def html_new(file, data):
    f = open(file, 'w')
    f.write(data)
    f.close()

#------------------------------------------------------------------------------
    
def html_append(file, data):
    f = open(file, 'a')
    f.write(data)
    f.close()

#------------------------------------------------------------------------------
    
def get_events(file, person):
    wb = openpyxl.load_workbook(file)    
    sheet = wb.get_sheet_by_name(person)
    
    events = []
    
    for cellObj in sheet.columns[0]:
        events.append(cellObj.value)
        
    events.remove('Event')
    
    return(events)
    
#------------------------------------------------------------------------------    
    
def get_people(file):
    wb = openpyxl.load_workbook(file)
    people = wb.get_sheet_names()
    return(people)

#------------------------------------------------------------------------------
    
def get_date(file, person, event):
    
    wb = openpyxl.load_workbook(file)    
    sheet = wb.get_sheet_by_name(person)    
    
    day = ''
    month = ''
    year = ''    
    
    for rowNum in range(2, sheet.get_highest_row()+1):  # skip the first row
        currentEvent = sheet.cell(row=rowNum, column = 1).value
        if currentEvent == event:
            day = sheet.cell(row=rowNum, column=2).value
            month = sheet.cell(row=rowNum, column=3).value
            year = sheet.cell(row=rowNum, column=4).value
    
    date = [day, month, year]
    return(date)    

#------------------------------------------------------------------------------

def get_description(file, person, event):
    wb = openpyxl.load_workbook(file)    
    sheet = wb.get_sheet_by_name(person)
    
    description = ''
    
    for rowNum in range(2, sheet.get_highest_row()+1):  # skip the first row
        currentEvent = sheet.cell(row=rowNum, column = 1).value
        if currentEvent == event:
            description = sheet.cell(row=rowNum, column=11).value
            
    return(description)

#------------------------------------------------------------------------------

def get_photo(file, person, event):
    wb = openpyxl.load_workbook(file)    
    sheet = wb.get_sheet_by_name(person)
    
    photo = ''
    
    for rowNum in range(2, sheet.get_highest_row()+1):  # skip the first row
        currentEvent = sheet.cell(row=rowNum, column = 1).value
        if currentEvent == event:
            photo = sheet.cell(row=rowNum, column=13).value
            
    return(photo)


#------------------------------------------------------------------------------
    
def get_year(file, person, event):
    
    wb = openpyxl.load_workbook(file)    
    sheet = wb.get_sheet_by_name(person)    
    
    year = ''    
    
    for rowNum in range(2, sheet.get_highest_row()+1):  # skip the first row
        currentEvent = sheet.cell(row=rowNum, column = 1).value
#        print(sheet.get_highest_row())
#        print('current event is' + currentEvent)
#        print('event is' + event)
        if currentEvent == event:
            year = sheet.cell(row=rowNum, column=4).value
    
    return(str(year))    

#------------------------------------------------------------------------------
    
def get_location(file, person, event):
    
    wb = openpyxl.load_workbook(file)    
    sheet = wb.get_sheet_by_name(person)    
    
    city = ''
    state = ''
    other = ''    
    
    for rowNum in range(1, sheet.get_highest_row()+1):  # skip the first row
        currentEvent = sheet.cell(row=rowNum, column = 1).value
        if currentEvent == event:
            city = sheet.cell(row=rowNum, column=8).value
            state = sheet.cell(row=rowNum, column=9).value
            other = sheet.cell(row=rowNum, column=10).value
    
    location = [city, state, other]
    return(location)    

#------------------------------------------------------------------------------    
    
def build_main(htmlFile, excelFile, person):
    html_new(htmlFile, start_doc(person))

    events = get_events(excelFile, person)
    links = {}
    
    for ev in events:
        newLink = person+'-'+ev+'.html'
        html_append(htmlFile, body_doc(get_year(excelFile,person,ev)+ ' ' + ev,newLink))
        links[ev] = newLink
        
    html_append(htmlFile, end_doc())
    
    build_events(links, excelFile, person)    
    
    return(links)

#------------------------------------------------------------------------------

def build_events(links, excelFile, person):
    
    for events in links:
        
        date = get_date(excelFile, person, events) 
        location = get_location(excelFile, person, events)
        description = get_description(excelFile, person, events)
        photo = get_photo(excelFile, person, events)
        
        html_new(links[events], start_doc(person))
        
        html_append(links[events], event_doc(events))
        if date[2]:
            html_append(links[events], date_doc(date[0],date[1],date[2]))
        if location[1]:
            html_append(links[events], where_doc(location[0],location[1]))
        if location[2]:
            html_append(links[events], where_other_doc(location[2]))
        if description:
            html_append(links[events], description_doc(description))
        if photo:
            html_append(links[events], photo_doc(photo))
        
        html_append(links[events], end_doc())

#------------------------------------------------------------------------------

def build_from_excel(file):
    wb = openpyxl.load_workbook(file)
    people = wb.get_sheet_names()
    
    build_index(file)    
    
    for names in people:
        build_main(names+'.html', file, names)
#------------------------------------------------------------------------------

def build_index(file):
    wb = openpyxl.load_workbook(file)
    people = wb.get_sheet_names()
    
    html_new('index.html', start_index_doc('Family Name...'))
    
    for names in people:
        html_append('index.html',index_doc(names))
        
    html_append('index.html',end_doc())
#------------------------------------------------------------------------------

if __name__ == '__main__':
    
    
    filename = askopenfilename()
    build_from_excel(filename)